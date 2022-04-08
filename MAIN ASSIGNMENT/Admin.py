#importing sys module
import sys
#importing openpyxl- Python library to read/write Excel
import openpyxl
#
def workbook(filename,sheetname):
   #using openpyxl.load_workbook() function to access Excel file
    wb = openpyxl.load_workbook(filename)
    #giving sheet,row,colum
    #to know the active sheet:sheet=wb.active
    sheet1 = wb[sheetname]
#determing the size of sheet with worksheet object's max_row & max_column attributes
    row = sheet1.max_row
    col = sheet1.max_column
    #print(row)
    #print(col)
    return row, col, sheet1,wb

#creating Admin() class so as to perform functionalities
class Admin():
    def __init__(self,adminName,adminPassword):
        self.adminName=adminName
        self.adminPassword=adminPassword
        print("******Welcome Admin*******\n 1. Add New movie\n 2. Edit Movie\n 3. Delete Movie\n 4.Logout")
    def adminFunctionalities(self, choice):
        #using while for all TRUE cases
        while(True):
            #Movies.xlsx is excel file having Movies_info
            row, col, sheet1, wb = workbook("Movies.xlsx", "Sheet1")
            #if choosen option is "1" then admin can Enter movie_details
            if(choice==1):
                print("Enter movie Details:")
                row1=1
                col1=1
                for a in range(1,col+1):
                    b=(sheet1.cell(row1,col1).value)
                    sheet1.cell(row+1,a,value=input(f" Enter {b}:"))
                    col1+=1
                    #new file "revised.xlsx" will be created which contains added movie_details
                wb.save("revised.xlsx")
                break
            elif(choice==2):
                #if choosen is "2" then admin can edit some fields ny giving movie name
                editMovie=input("Enter the movie name to be Edited: ")
                #col1=1
                for i in range(2,row+1):
                    name=sheet1.cell(i,column=1).value
                    #checking whether movie_name entered is available so as to perform edits
                    if(editMovie==name):
                        for j in range(1,col+1):
                               sheet1.cell(i,j,value=input(f"Enter the {sheet1.cell(1,j).value} to change:"))
                #Same file "Movies.xlsx" will be updated after performing required edits by Admin
                wb.save("Movies.xlsx")
                break
            elif(choice==3):
                deleteMovie=input("Enter name of movie to be deleted:")
                for i in range(2,row+1):
                    name=sheet1.cell(i,column=1).value
                    if(name==deleteMovie):
                        sheet1.delete_rows(i,True)
                #new file "Revised2.xlsx" is created after performing the required action
                wb.save("Revised2.xlsx")
                break
                #If "4" is choosen then Admin can LOG_OUT Successfully
            elif(choice==4):
                print("Logged out Succesfully !!")
                sys.exit()

#creating "user_credentials" class so as to perform USER_FUNCTIONALITIES
class UserCredentials(Admin):
    def __init__(self,userName,userPassword):
        #storing the user_credentials into dictionary
        self.userCredentials={userName:userPassword}
        print(f"***** Hey {self.userCredentials},you have Logged in Successfully *****")
     #user_registration
    def validateUser(self, username, password):
        if username in self.userCredentials.keys() and password == self.userCredentials[username]:
            print("Logged-in Successfully")
        else:
            print("Sorry,Login Failed!! ") #Incorrect Credentials

    def user_Functionalities(self):
        wb = openpyxl.load_workbook("MoviesInfo.xlsx")
        sh1 = wb['Sheet1']
        row = sh1.max_row
        col = sh1.max_column
        m = 1
        for i in range(2,row+1):
            nextMovie=print(sh1.cell(i,1).value)
            print(f"{nextMovie}")
            m+=1

    def Booktickets(self, Userchoice):
        row, col, sh1, wb = workbook("Movies.xlsx", "Sheet1")
        for i in range(1, col + 1):
            print(sh1.cell(1, i).value, end=": ")
            print(sh1.cell(Userchoice + 1, i).value)
        while (True):
            print("1.Book Tickets\n 2. Cancel Ticket\n 3. Give User Rating\n 4.exit")
            TicketChoice = int(input("Enter the Choice: "))
            if (TicketChoice == 1):
                print("***** Welcome User *****")
                for i in range(7, 9):
                    print(f"Timing : {sh1.cell(Userchoice + 1, i).value}")
                timing = int(input("enter the timing you want to choose: "))
                print(f"Selected timing is {timing}")
                print(f"No of Remaining Seats are {sh1.cell(Userchoice + 1, 13).value}")
                totalTicketsToBeBooked = int(input("Enter the number of tickets to be Booked:"))
                ticketsAvailable = sh1.cell(Userchoice + 1, 13).value
                sh1.cell(Userchoice + 1, 13, value=ticketsAvailable - totalTicketsToBeBooked)
                print(sh1.cell(Userchoice + 1, 13).value)
                print("***** Your Tickets Booked *****")
                wb.save("Movies.xlsx")
            elif (TicketChoice == 2):
                ticketsTobeCancelled = int(input("enter the Number of tickets to be Cancelled"))
                ticketsAvailable = sh1.cell(Userchoice + 1, 13).value
                sh1.cell(Userchoice + 1, 13, value=ticketsAvailable + ticketsTobeCancelled)
                print(sh1.cell(Userchoice + 1, 13).value)
                wb.save("Movies.xlsx")
            elif (TicketChoice == 3):
                userRating = input("Kindly provide your Review for above movie")
            elif (TicketChoice == 4):
                break
    def moviesInfo(self):
        row, col, sh1, wb = workbook("Movies.xlsx", "Sheet1")
        b = 1
        for i in range(2, row + 1):
            print(b, end='.')
            print(sh1.cell(i, 1).value)
            b += 1
        print("Logout")

        Userchoice = int(input("Enter your choose: "))
        if (Userchoice == 1):
            print("***** Welcome User *****")
            self.Booktickets(1)
        elif (Userchoice == 2):
            print("***** Welcome User *****")
            self.Booktickets(2)
        elif (Userchoice == 3):
            print("***** Welcome User *****")
            self.Booktickets(3)
while(True):
    print("***** Welcome to BookMyShow *****\n  1.Login\n  2.Register\n  3.Exit")
    user_inp=int(input("Enter your choice:"))
    if(user_inp==3):
        break
    elif(user_inp==1):
        print("Welcome\n 1.Admin Login \n2.UserLogin")
        choice=int(input("Enter your Choice:"))
        if(choice==1):
            name=input("Enter your Name:")
            password=input("Enter your Password: ")
            admin=Admin(name,password)
            admin.adminFunctionalities()

        elif(choice==2):
            userName=input("enter user_name:")
            userPassword=input("enter password:")
            SaiKumar=UserCredentials(userName,userPassword)
            SaiKumar.validateUser(userName,userPassword)
            SaiKumar.moviesInfo()
    elif(user_inp==2):
        row, col, sh1, wb = workbook("UserRegisteredDetails.xlsx", "Sheet1")
        for i in range(1, col + 1):
            sh1.cell(row + 1, i, value=input(f"Enter your {sh1.cell(1, i).value}"))
        wb.save("UserRegisteredDetails.xlsx")
    else:
        sys.exit()
